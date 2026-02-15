"""
Excel playbook writer with topic-based sheet organization.

Creates professional contract playbooks matching the structure of
high-quality legal playbooks with separate sheets per topic.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Styling
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
TITLE_FONT = Font(bold=True, size=18, color="2B579A")
SECTION_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical='top')
ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# Column structure for clause analysis sheets (header names only, widths are auto-calculated)
CLAUSE_COLUMNS = [
    "Section",
    "Subsection",
    "Issue",
    "Current Language",
    "Purpose/Rationale",
    "Customer Concerns",
    "Customer Edits to Watch",
    "Provider Position",
    "Acceptable Modifications",
    "Fallback Language",
    "Do Not Accept",
    "Notes",
]

# Min/max constraints for column widths (in characters)
MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 65

# Row height settings
MIN_ROW_HEIGHT = 20
MAX_ROW_HEIGHT = 200
CHARS_PER_LINE = 60  # Approximate characters per line at typical column width
LINE_HEIGHT_PTS = 15  # Points per line of text


def auto_fit_columns(ws, min_width=MIN_COL_WIDTH, max_width=MAX_COL_WIDTH):
    """Auto-fit column widths based on the longest content in each column."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                # For multi-line content, use the longest single line
                text = str(cell.value)
                lines = text.split('\n')
                longest_line = max(len(line) for line in lines) if lines else 0
                # Account for bold/header fonts being slightly wider
                if cell.font and cell.font.bold:
                    longest_line = int(longest_line * 1.1)
                max_len = max(max_len, longest_line)
        # Add padding, then clamp between min and max
        width = min(max(max_len + 4, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def auto_fit_row_heights(ws, start_row=1, col_widths=None):
    """Auto-fit row heights based on content and the column width it wraps into."""
    for row in ws.iter_rows(min_row=start_row):
        max_lines = 1
        for cell in row:
            if cell.value is not None:
                text = str(cell.value)
                col_letter = get_column_letter(cell.column)
                col_width = col_widths.get(col_letter, MAX_COL_WIDTH) if col_widths else MAX_COL_WIDTH
                # Approximate chars that fit per line in this column
                chars_per_line = max(int(col_width * 1.2), 10)
                # Count explicit newlines plus wrapped lines
                lines = text.split('\n')
                total_lines = 0
                for line in lines:
                    wrapped = max(1, -(-len(line) // chars_per_line))  # Ceiling division
                    total_lines += wrapped
                max_lines = max(max_lines, total_lines)
        height = max(MIN_ROW_HEIGHT, min(max_lines * LINE_HEIGHT_PTS, MAX_ROW_HEIGHT))
        ws.row_dimensions[row[0].row].height = height


def auto_size_sheet(ws, start_row=1):
    """Apply auto-fit to both columns and rows for a worksheet."""
    auto_fit_columns(ws)
    # Build a col_widths dict for row height calculation
    col_widths = {}
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        col_widths[col_letter] = ws.column_dimensions[col_letter].width
    auto_fit_row_heights(ws, start_row=start_row, col_widths=col_widths)


def generate_playbook_excel(playbook_data: dict, output_path: str):
    """
    Generate a professional Excel playbook with topic-based sheets.

    Args:
        playbook_data: The structured playbook data from Claude analysis
        output_path: Path to save the Excel file
    """
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create Overview sheet
    create_overview_sheet(wb, playbook_data.get("overview", {}))

    # Create topic sheets
    topics = playbook_data.get("topics", {})
    for topic_name, clauses in topics.items():
        if clauses:  # Only create sheet if there are clauses
            create_topic_sheet(wb, topic_name, clauses)

    # Create Quick Reference sheet
    create_quick_reference_sheet(wb, playbook_data.get("quick_reference", []))

    wb.save(output_path)
    return output_path


def create_overview_sheet(wb: Workbook, overview: dict):
    """Create the Overview sheet with agreement summary and guidance."""
    ws = wb.create_sheet("Overview", 0)

    # Title
    title = overview.get("title", "Contract Playbook")
    ws["A1"] = f"{title} Contracting Playbook"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    row = 3

    # Agreement details
    details = [
        ("Agreement Type:", overview.get("agreement_type", "")),
        ("Perspective:", overview.get("perspective", "")),
        ("Parties:", overview.get("parties", "")),
        ("Effective Date:", overview.get("effective_date", "")),
        ("Governing Law:", overview.get("governing_law", "")),
    ]

    for label, value in details:
        if value:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

    row += 1

    # Key Principles
    ws.cell(row=row, column=1, value="KEY PRINCIPLES").font = SECTION_FONT
    row += 1

    for i, principle in enumerate(overview.get("key_principles", []), 1):
        ws.cell(row=row, column=1, value=f"{i}. {principle}")
        ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        row += 1

    row += 1

    # Executive Summary
    ws.cell(row=row, column=1, value="EXECUTIVE SUMMARY").font = SECTION_FONT
    row += 1
    summary = overview.get("executive_summary", "")
    ws.cell(row=row, column=1, value=summary)
    ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
    ws.merge_cells(f"A{row}:B{row}")
    ws.row_dimensions[row].height = 100
    row += 2

    # How to Use
    ws.cell(row=row, column=1, value="HOW TO USE THIS PLAYBOOK").font = SECTION_FONT
    row += 1

    for i, instruction in enumerate(overview.get("how_to_use", []), 1):
        ws.cell(row=row, column=1, value=f"{i}. {instruction}")
        ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        row += 1

    # Auto-fit columns and rows
    auto_fit_columns(ws, min_width=25, max_width=90)
    # Ensure column B is wide enough for long content
    if ws.column_dimensions["B"].width < 60:
        ws.column_dimensions["B"].width = 60
    auto_fit_row_heights(ws, start_row=1, col_widths={
        "A": ws.column_dimensions["A"].width,
        "B": ws.column_dimensions["B"].width,
    })


def create_topic_sheet(wb: Workbook, topic_name: str, clauses: list):
    """Create a sheet for a specific contract topic."""
    # Sanitize sheet name - remove invalid characters and truncate
    # Excel doesn't allow: / \ ? * [ ] :
    sheet_name = topic_name.replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "").replace(":", "-")
    sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
    ws = wb.create_sheet(sheet_name)

    # Headers
    for col_idx, header in enumerate(CLAUSE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add clause data
    for row_idx, clause in enumerate(clauses, 2):
        ws.cell(row=row_idx, column=1, value=clause.get("section", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=clause.get("subsection", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=clause.get("issue", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=clause.get("current_language", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=clause.get("purpose_rationale", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=6, value=clause.get("customer_concerns", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=7, value=clause.get("customer_edits_to_watch", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=8, value=clause.get("provider_position", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=9, value=clause.get("acceptable_modifications", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=10, value=clause.get("fallback_language", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=11, value=clause.get("do_not_accept", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=12, value=clause.get("notes", "")).border = THIN_BORDER

        # Apply wrap text and alternating row colors
        for col in range(1, 13):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = WRAP_ALIGNMENT
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto-fit columns and rows based on content (skip header row for row heights)
    auto_size_sheet(ws, start_row=2)


def create_quick_reference_sheet(wb: Workbook, quick_reference: list):
    """Create the Quick Reference sheet with hard limits."""
    ws = wb.create_sheet("Quick Reference")

    # Title
    ws["A1"] = "Quick Reference - Hard Limits"
    ws["A1"].font = Font(bold=True, size=14, color="2B579A")
    ws.merge_cells("A1:B1")

    ws["A2"] = "Items below require executive approval before deviating from standard position"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:B2")

    # Headers
    ws["A4"] = "Topic"
    ws["B4"] = "Hard Limit (Do Not Accept Without Executive Approval)"
    ws["A4"].font = HEADER_FONT
    ws["B4"].font = HEADER_FONT
    ws["A4"].fill = HEADER_FILL
    ws["B4"].fill = HEADER_FILL
    ws["A4"].border = THIN_BORDER
    ws["B4"].border = THIN_BORDER

    # Data
    for row_idx, item in enumerate(quick_reference, 5):
        issue_cell = ws.cell(row=row_idx, column=1, value=item.get("issue", ""))
        limit_cell = ws.cell(row=row_idx, column=2, value=item.get("limit", ""))
        issue_cell.alignment = WRAP_ALIGNMENT
        limit_cell.alignment = WRAP_ALIGNMENT
        issue_cell.border = THIN_BORDER
        limit_cell.border = THIN_BORDER
        if row_idx % 2 == 0:
            issue_cell.fill = ALT_ROW_FILL
            limit_cell.fill = ALT_ROW_FILL

    # Auto-fit columns and rows
    auto_fit_columns(ws, min_width=25, max_width=80)
    if ws.column_dimensions["B"].width < 60:
        ws.column_dimensions["B"].width = 60
    auto_fit_row_heights(ws, start_row=5, col_widths={
        "A": ws.column_dimensions["A"].width,
        "B": ws.column_dimensions["B"].width,
    })

    # Freeze header
    ws.freeze_panes = "A5"
